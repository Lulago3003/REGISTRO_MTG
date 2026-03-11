import json, sys, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── USAGE ─────────────────────────────────────────────────────────
# python3 json_to_excel.py  MTG_Backup_2024-11-20.json
# Output: MTG_PAOER_2024-11-20.xlsx  (same folder)
# ──────────────────────────────────────────────────────────────────

if len(sys.argv) < 2:
    print("Uso: python3 json_to_excel.py  archivo_backup.json")
    sys.exit(1)

json_path = Path(sys.argv[1])
with open(json_path, encoding="utf-8") as f:
    properties = json.load(f)

print(f"✅ {len(properties)} propiedades cargadas...")

# ── PALETTE ───────────────────────────────────────────────────────
GREEN_DARK  = "1A6E45"
GREEN_MID   = "228B55"
GREEN_LIGHT = "4CB87A"
GREEN_PALE  = "E2F5EB"
GREEN_PALE2 = "F0FAF4"
WHITE       = "FFFFFF"
GRAY_DARK   = "1A2E22"
GRAY_MID    = "5A7A6A"
GRAY_LIGHT  = "C8E6D5"
YELLOW_PALE = "FFFBEA"
YELLOW_BDR  = "E8DDA0"
RED_PALE    = "FDF0F0"

def F(hex): return PatternFill("solid", fgColor=hex)
def ft(bold=False, size=10, color="1A2E22", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(color="C8E6D5", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def bdr_header():
    m = Side(style="medium", color=GREEN_DARK)
    t = Side(style="thin",   color=GREEN_DARK)
    return Border(left=t, right=t, top=m, bottom=m)

wb = Workbook()

# ════════════════════════════════════════════════════════════════
def make_sheet(wb, title_text, col_defs, data_rows, mode="normal"):
    """
    mode = "normal" | "invoice" | "circuits"
    col_defs = list of (header_label, col_width)
    data_rows = list of lists
    """
    ws = wb.create_sheet(title_text)
    ws.sheet_view.showGridLines = False

    ncols = len(col_defs)
    last_col = get_column_letter(ncols)

    # Column widths
    for i, (_, w) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ROW 1: top padding ──────────────────────────────────────
    ws.row_dimensions[1].height = 6
    for c in range(1, ncols+1):
        ws.cell(1, c).fill = F(GREEN_DARK)

    # ── ROW 2: BIG TITLE ────────────────────────────────────────
    ws.row_dimensions[2].height = 46
    ws.merge_cells(f"A2:{last_col}2")
    tc = ws.cell(2, 1, f"  META TECHNOLOGY GLOBAL  \u2014  {title_text}")
    tc.font      = ft(bold=True, size=16, color=WHITE)
    tc.fill      = F(GREEN_DARK)
    tc.alignment = al("left", "center")
    for c in range(2, ncols+1):
        ws.cell(2, c).fill = F(GREEN_DARK)

    # ── ROW 3: sub-banner ───────────────────────────────────────
    ws.row_dimensions[3].height = 5
    for c in range(1, ncols+1):
        ws.cell(3, c).fill = F(GREEN_MID)

    # ── ROW 4: COLUMN HEADERS ───────────────────────────────────
    ws.row_dimensions[4].height = 36
    for i, (label, _) in enumerate(col_defs, 1):
        c = ws.cell(4, i, label.upper())
        c.font      = ft(bold=True, size=9, color=WHITE)
        c.fill      = F(GREEN_MID)
        c.alignment = al("center", "center", wrap=True)
        c.border    = bdr_header()

    ws.freeze_panes = "A5"

    # ── DATA ROWS ────────────────────────────────────────────────
    for ri, row in enumerate(data_rows):
        excel_row = 5 + ri
        even = ri % 2 == 0
        ws.row_dimensions[excel_row].height = 22

        if mode == "invoice":
            row_bg  = YELLOW_PALE if even else WHITE
            bdr_col = YELLOW_BDR
        else:
            row_bg  = GREEN_PALE2 if even else WHITE
            bdr_col = "D4EDE0"

        for ci, val in enumerate(row):
            c = ws.cell(excel_row, ci+1, val)
            c.fill      = F(row_bg)
            c.font      = ft(size=10, color=GRAY_DARK)
            c.alignment = al("left", "center", wrap=(ci == len(row)-1))
            c.border    = bdr(bdr_col)

            # Money columns in invoice
            if mode == "invoice" and ci in (5, 7):
                c.number_format = '"$"#,##0.00'
                c.font = ft(bold=True, size=10, color=GREEN_DARK)
                c.fill = F(GREEN_PALE)
                c.alignment = al("center","center")
            # Average column
            if mode == "invoice" and ci == 8:
                c.number_format = '"$"#,##0.00'
                c.font = ft(bold=True, size=11, color=GREEN_DARK)
                c.fill = F(GREEN_PALE)
                c.alignment = al("center","center")
                c.border = bdr(GREEN_LIGHT, "medium")
            # N° column
            if ci == 0:
                c.font      = ft(bold=True, size=10, color=GREEN_MID)
                c.fill      = F(GREEN_PALE)
                c.alignment = al("center","center")
                c.border    = bdr(GREEN_LIGHT)
            # Breaker number highlight
            if mode == "circuits" and ci == 3:
                c.font      = ft(bold=True, size=10, color=WHITE)
                c.fill      = F(GREEN_MID)
                c.alignment = al("center","center")
            # Estado eléctrico color
            if mode == "normal" and ci == 21:
                colors = {"Excelente":GREEN_PALE,"Bueno":"FFFDE8","Regular":"FEF3E8","Deficiente":RED_PALE}
                text_c = {"Excelente":GREEN_DARK,"Bueno":"7A6000","Regular":"C96A1A","Deficiente":"B83232"}
                v = str(val)
                c.fill = F(colors.get(v, WHITE))
                c.font = ft(bold=True, size=10, color=text_c.get(v, GRAY_DARK))

    # ── PRINT SETUP ─────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = "4:4"

    return ws

# ════════════════════════════════════════════════════════════════
#  PORTADA
# ════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Portada"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 30
ws0.column_dimensions["C"].width = 36
ws0.column_dimensions["D"].width = 20
ws0.column_dimensions["E"].width = 20
ws0.column_dimensions["F"].width = 3

# ── Banner rows ─────────────────────────────────────────────────
ws0.row_dimensions[1].height = 6
for c in range(1,7): ws0.cell(1,c).fill = F(GREEN_DARK)

ws0.row_dimensions[2].height = 52
for c in range(1,7): ws0.cell(2,c).fill = F(GREEN_DARK)

ws0.row_dimensions[3].height = 28
for c in range(1,7): ws0.cell(3,c).fill = F(GREEN_DARK)

ws0.row_dimensions[4].height = 5
for c in range(1,7): ws0.cell(4,c).fill = F(GREEN_MID)

ws0.row_dimensions[5].height = 14
for c in range(1,7): ws0.cell(5,c).fill = F(GREEN_PALE)

ws0.merge_cells("B2:E2")
t = ws0["B2"]
t.value = "  META TECHNOLOGY GLOBAL"
t.font  = ft(bold=True, size=24, color=WHITE)
t.fill  = F(GREEN_DARK)
t.alignment = al("left","center")

ws0.merge_cells("B3:E3")
s = ws0["B3"]
s.value = "  Registro de Instalaciones    \u00b7    Plataforma PAOER    \u00b7    Emporia Vue 3"
s.font  = ft(italic=True, size=11, color="A8D4BC")
s.fill  = F(GREEN_DARK)
s.alignment = al("left","center")

# ── Info rows ───────────────────────────────────────────────────
def cover_row(row, label, val, bg_l=GREEN_PALE2, bg_v=WHITE):
    ws0.row_dimensions[row].height = 28
    lc = ws0.cell(row, 2, f"  {label}")
    lc.font=ft(bold=True,size=10,color=GRAY_MID); lc.fill=F(bg_l)
    lc.alignment=al("left","center")
    lc.border=bdr(GRAY_LIGHT)
    ws0.merge_cells(f"C{row}:E{row}")
    vc=ws0.cell(row,3,f"  {val}")
    vc.font=ft(bold=True,size=10,color=GREEN_DARK); vc.fill=F(bg_v)
    vc.alignment=al("left","center"); vc.border=bdr(GRAY_LIGHT)

cover_row(6,  "Fecha de Exportaci\u00f3n",              datetime.date.today().strftime("%d/%m/%Y"),  GREEN_PALE2, WHITE)
cover_row(7,  "Total de Propiedades",                    str(len(properties)),                        WHITE,       GREEN_PALE2)
total_bk = sum(len(p.get("breakers",[])) for p in properties)
cover_row(8,  "Total de Circuitos Registrados",          str(total_bk),                               GREEN_PALE2, WHITE)
solar = sum(1 for p in properties if p.get("energyType","")!="Solo Red El\u00e9ctrica")
cover_row(9,  "Propiedades con Sistema Solar/Bater\u00eda", str(solar),                              WHITE,       GREEN_PALE2)
cover_row(10, "Empresa",                                 "Meta Technology Global",                    GREEN_PALE2, WHITE)

# ── Blank spacer ────────────────────────────────────────────────
ws0.row_dimensions[11].height = 14
for c in range(1,7): ws0.cell(11,c).fill = F(GREEN_PALE)

# ── Sheet index header ───────────────────────────────────────────
ws0.row_dimensions[12].height = 26
ws0.merge_cells("B12:E12")
hi=ws0["B12"]; hi.value="  Contenido del Archivo"
hi.font=ft(bold=True,size=11,color=WHITE); hi.fill=F(GREEN_MID); hi.alignment=al("left","center")
hi.border=bdr(GREEN_DARK)

sheets_info=[
  ("  \u25ba  Propiedades","Datos completos de cada instalaci\u00f3n registrada"),
  ("  \u25ba  Circuitos",  "Mapa de breakers, sensores CT y amperajes"),
  ("  \u25ba  Facturas",   "Facturas el\u00e9ctricas reportadas por el cliente"),
]
for i,(nm,desc) in enumerate(sheets_info):
    r=13+i; ws0.row_dimensions[r].height=24
    nc=ws0.cell(r,2,nm); nc.font=ft(bold=True,size=10,color=GREEN_DARK); nc.fill=F(GREEN_PALE); nc.alignment=al("left","center"); nc.border=bdr(GREEN_LIGHT)
    ws0.merge_cells(f"C{r}:E{r}")
    dc=ws0.cell(r,3,f"  {desc}"); dc.font=ft(size=10,color=GRAY_MID); dc.fill=F(WHITE); dc.alignment=al("left","center"); dc.border=bdr(GREEN_LIGHT)

# ════════════════════════════════════════════════════════════════
#  PROPIEDADES
# ════════════════════════════════════════════════════════════════
prop_cols = [
    ("N°",22,5),("Nombre / ID",None,26),("Tipo",None,14),("Propietario",None,22),
    ("Teléfono",None,16),("Dirección",None,36),("Habitaciones",None,13),
    ("Fase Eléctrica",None,18),("Sistema Energético",None,22),
    ("kWp Solar",None,11),("N° Paneles",None,11),("Inversor",None,20),
    ("Cap. Batería",None,13),("Modelo Batería",None,20),
    ("Técnico",None,22),("Fecha Instalación",None,17),
    ("Serie Emporia",None,20),("Ubicación Medidor",None,22),
    ("Canales CT",None,11),("Conectividad",None,14),("WiFi SSID",None,18),
    ("Estado Eléctrico",None,18),("Obs. Medidor",None,26),
    ("N° Circuitos",None,12),("Observaciones",None,34),
]
prop_defs = [(h,w) for h,_,w in prop_cols]
prop_rows = []
for i,p in enumerate(properties):
    ee=p.get("energyExtra",{})
    prop_rows.append([
        i+1, p.get("name",""), p.get("propType",""), p.get("owner",""),
        p.get("phone",""), p.get("address",""), p.get("rooms",""),
        p.get("phase",""), p.get("energyType",""),
        ee.get("solarKw",""), ee.get("solarPanels",""), ee.get("inverter",""),
        ee.get("batCap",""), ee.get("batModel",""),
        p.get("tech",""), p.get("installDate",""),
        p.get("serial",""), p.get("meterLoc",""),
        p.get("ctChannels",""), p.get("connectivity",""), p.get("wifi",""),
        p.get("elecState",""), p.get("meterObs",""),
        len(p.get("breakers",[])), p.get("notes","")
    ])
make_sheet(wb, "Propiedades", prop_defs, prop_rows, "normal")

# ════════════════════════════════════════════════════════════════
#  CIRCUITOS
# ════════════════════════════════════════════════════════════════
circ_defs=[("N°",5),("Propiedad",26),("Dirección",32),("N° Breaker",11),
           ("Circuito",22),("Tipo de Carga",22),("Sensor CT",16),("Amperaje",12),("Observaciones",34)]
circ_rows=[]; cn=0
for p in properties:
    for b in p.get("breakers",[]):
        cn+=1
        circ_rows.append([cn,p.get("name",""),p.get("address",""),
                          b.get("number",""),b.get("name",""),b.get("type",""),
                          b.get("sensor",""),b.get("amp",""),b.get("obs","")])
make_sheet(wb, "Circuitos", circ_defs, circ_rows, "circuits")

# ════════════════════════════════════════════════════════════════
#  FACTURAS
# ════════════════════════════════════════════════════════════════
fact_defs=[("N°",5),("Propiedad",26),("Propietario",22),("Dirección",32),
           ("Mes Reciente",16),("Factura Reciente ($)",20),
           ("Mes Anterior",16),("Factura Anterior ($)",20),
           ("Promedio ($)",16),("Observaciones",30)]
fact_rows=[]; fn=0
for p in properties:
    if p.get("bill1") or p.get("bill2"):
        fn+=1
        b1=float(p["bill1"]) if p.get("bill1") else 0
        b2=float(p["bill2"]) if p.get("bill2") else 0
        avg=(b1+b2)/2 if b1 and b2 else (b1 or b2)
        fact_rows.append([fn, p.get("name",""), p.get("owner",""), p.get("address",""),
                          p.get("bill1Month",""), b1, p.get("bill2Month",""), b2, avg, ""])
make_sheet(wb, "Facturas", fact_defs, fact_rows, "invoice")

# ── SAVE ──────────────────────────────────────────────────────
out = json_path.parent / f"MTG_PAOER_{datetime.date.today().isoformat()}.xlsx"
wb.save(out)
print(f"✅ Excel generado: {out}")
print(f"   Propiedades : {len(properties)}")
print(f"   Circuitos   : {cn}")
print(f"   Con facturas: {fn}")
